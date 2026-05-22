from flask import Flask, render_template_string, request, redirect, jsonify
import sqlite3
import datetime
from io import BytesIO
from urllib.parse import urlencode

from openpyxl import Workbook
from openpyxl.styles import Font

app = Flask(__name__)

PER_PAGE = 30

CSS = """
<style>
:root {
    --bg: #f4f6fb;
    --surface: #ffffff;
    --surface-soft: #f8faff;
    --text: #182033;
    --muted: #667085;
    --border: #e6eaf2;
    --primary: #4f46e5;
    --primary-hover: #4338ca;
    --danger: #ef4444;
    --danger-hover: #dc2626;
    --success: #22c55e;
    --warning: #f59e0b;
    --info: #3b82f6;
    --gray: #64748b;
    --gray-hover: #475569;
    --radius: 14px;
    --shadow: 0 10px 30px rgba(15, 23, 42, 0.08);
}
* { box-sizing: border-box; }
body {
    margin: 0;
    background:
        radial-gradient(circle at top right, rgba(79, 70, 229, .16) 0%, rgba(79, 70, 229, 0) 30%),
        linear-gradient(180deg, #eef2ff 0%, #f8fafc 42%, #f4f6fb 100%);
    color: var(--text);
    font-family: "Inter", "Segoe UI", Arial, sans-serif;
    line-height: 1.45;
}
.navbar {
    position: sticky;
    top: 0;
    z-index: 10;
    background: linear-gradient(135deg, #111827 0%, #1e1b4b 52%, #312e81 100%);
    border-bottom: 1px solid rgba(255, 255, 255, .14);
    box-shadow: 0 12px 30px rgba(15, 23, 42, .28);
}
.navbar-inner {
    max-width: 1240px;
    margin: 0 auto;
    padding: 14px 20px;
    display: flex;
    align-items: center;
    gap: 12px;
    flex-wrap: wrap;
}
.nav-brand {
    display: inline-flex;
    align-items: center;
    justify-content: center;
    margin-right: 8px;
}
.georgia-flag {
    position: relative;
    width: 60px;
    height: 40px;
    border-radius: 6px;
    background: #fff;
    border: 1px solid rgba(255, 255, 255, .85);
    box-shadow: 0 8px 22px rgba(0, 0, 0, .24), inset 0 0 0 1px rgba(220, 38, 38, .08);
    overflow: hidden;
}
.georgia-flag::before,
.georgia-flag::after {
    content: "";
    position: absolute;
    left: 50%;
    top: 50%;
    transform: translate(-50%, -50%);
    background: #dc2626;
}
.georgia-flag::before {
    width: 6px;
    height: 100%;
}
.georgia-flag::after {
    width: 100%;
    height: 6px;
}
.georgia-flag span {
    position: absolute;
    z-index: 1;
    width: 7px;
    height: 7px;
    border-radius: 999px;
    background: #ef4444;
    transform: translate(-50%, -50%);
    box-shadow: 0 0 0 3px rgba(239, 68, 68, .16), 0 0 12px rgba(239, 68, 68, .85);
}
.georgia-flag .dot-1 { top: 25%; left: 25%; }
.georgia-flag .dot-2 { top: 25%; left: 75%; }
.georgia-flag .dot-3 { top: 75%; left: 25%; }
.georgia-flag .dot-4 { top: 75%; left: 75%; }
.nav-link {
    display: inline-flex;
    align-items: center;
    padding: 12px 18px;
    border-radius: 14px;
    border: 1px solid rgba(255, 255, 255, .18);
    background: rgba(255, 255, 255, .10);
    color: #f8fafc;
    font-size: 15px;
    font-weight: 800;
    text-decoration: none;
    box-shadow: inset 0 1px 0 rgba(255, 255, 255, .12);
    transition: all .2s ease;
}
.nav-link:hover {
    background: #f8fafc;
    color: #1e1b4b;
    border-color: #f8fafc;
    transform: translateY(-2px);
    box-shadow: 0 10px 24px rgba(15, 23, 42, .24);
}
.page {
    max-width: 1240px;
    margin: 0 auto;
    padding: 24px 20px 34px;
}
h1 {
    margin: 8px 0 20px;
    padding: 18px 22px;
    border-radius: 20px;
    background: linear-gradient(135deg, #111827 0%, #1e1b4b 58%, #4338ca 100%);
    color: #fff;
    font-size: 30px;
    letter-spacing: -0.02em;
    box-shadow: 0 18px 36px rgba(30, 27, 75, .18);
}
.panel, .table-wrap {
    border: 1px solid rgba(49, 46, 129, .10);
    border-radius: 18px;
    box-shadow: var(--shadow);
    background: linear-gradient(180deg, #ffffff 0%, #fbfcff 100%);
}
.panel {
    position: relative;
    padding: 20px;
    margin-bottom: 16px;
    overflow: hidden;
}
.panel::before {
    content: "";
    position: absolute;
    top: 0;
    left: 0;
    right: 0;
    height: 4px;
    background: linear-gradient(90deg, #111827, #4f46e5, #22c55e);
}
.section-title {
    display: inline-flex;
    margin: 0 0 14px;
    padding: 6px 10px;
    border-radius: 999px;
    background: #eef2ff;
    color: #312e81;
    font-size: 12px;
    font-weight: 800;
    text-transform: uppercase;
    letter-spacing: .06em;
}
.grid {
    display: grid;
    grid-template-columns: repeat(12, minmax(0, 1fr));
    gap: 14px;
}
.col-3 { grid-column: span 3; }
.col-4 { grid-column: span 4; }
.col-6 { grid-column: span 6; }
.col-8 { grid-column: span 8; }
.col-12 { grid-column: span 12; }
@media (max-width: 1000px) {
    .col-3, .col-4, .col-6, .col-8, .col-12 { grid-column: span 12; }
}
label { display:block; font-size:13px; font-weight:700; color:#1e1b4b; margin:0 0 6px; }
input, textarea, select {
    width: 100%;
    border: 1px solid #d7deeb;
    border-radius: 12px;
    padding: 10px 12px;
    font-size: 14px;
    color: var(--text);
    background: #fff;
    margin: 0;
    box-shadow: inset 0 1px 2px rgba(15, 23, 42, .03);
}
textarea { min-height: 96px; resize: vertical; }
input:focus, textarea:focus, select:focus {
    outline: none;
    border-color: #a5b4fc;
    box-shadow: 0 0 0 3px rgba(99, 102, 241, .15);
}
.date-compact { max-width: 170px; }
.actions { display:flex; flex-wrap:wrap; gap:10px; }
a.button, button {
    border: none;
    border-radius: 10px;
    padding: 9px 14px;
    text-decoration: none;
    cursor: pointer;
    font-size: 14px;
    font-weight: 600;
    background: var(--primary);
    color: #fff;
    transition: all .2s ease;
}
a.button:hover, button:hover { background: var(--primary-hover); transform: translateY(-1px); }
.button-red { background: var(--danger); }
.button-red:hover { background: var(--danger-hover); }
.button-gray { background: var(--gray); }
.button-gray:hover { background: var(--gray-hover); }
.button-status-accepted,
.compact-toolbar .button.button-status-accepted { background: var(--info); color: #fff; }
.button-status-work,
.compact-toolbar .button.button-status-work { background: var(--warning); color: #1f2937; }
.button-status-ready,
.compact-toolbar .button.button-status-ready { background: var(--success); color: #fff; }
.button-status-given,
.compact-toolbar .button.button-status-given { background: var(--gray); color: #fff; }
.button-status-accepted:hover,
.compact-toolbar .button.button-status-accepted:hover { background: #2563eb; }
.button-status-work:hover,
.compact-toolbar .button.button-status-work:hover { background: #d97706; color: #111827; }
.button-status-ready:hover,
.compact-toolbar .button.button-status-ready:hover { background: #16a34a; }
.button-status-given:hover,
.compact-toolbar .button.button-status-given:hover { background: var(--gray-hover); }
.table-wrap { overflow-x: auto; border-top: 4px solid #312e81; }
table { width: 100%; border-collapse: separate; border-spacing: 0; }
th, td { padding: 12px; border-bottom: 1px solid var(--border); text-align:left; vertical-align: top; }
th { background: #eef2ff; color: #1e1b4b; font-weight: 800; }
.orders-table th, .orders-table td { padding: 8px 9px; font-size: 13px; }
.orders-table .id-cell, .orders-table .price-cell { white-space: nowrap; width: 1%; }
.orders-table .phone-cell, .orders-table .date-cell { white-space: nowrap; }
.orders-table .device-cell { max-width: 140px; word-break: break-word; }
.orders-table .problem-cell { max-width: 210px; word-break: break-word; }
.orders-table .open-cell { width: 1%; white-space: nowrap; }
.orders-table .button { padding: 7px 10px; font-size: 13px; }
tr:last-child td { border-bottom: none; }
tr:hover td { background: #fbfcff; }
.status-tag { padding: 6px 10px; border-radius: 999px; font-size: 12px; font-weight: 700; color: #fff; display:inline-block; }
.status-accepted { background: var(--info); }
.status-work { background: var(--warning); color: #111827; }
.status-ready { background: var(--success); }
.status-given { background: #64748b; }
.status-select {
    min-width: 118px;
    width: auto;
    border: none;
    border-radius: 999px;
    padding: 6px 9px;
    font-size: 11px;
    font-weight: 700;
    color: #fff;
    cursor: pointer;
}
.status-select-accepted { background: var(--info); }
.status-select-work { background: var(--warning); color: #111827; }
.status-select-ready { background: var(--success); }
.status-select-given { background: var(--gray); }
.status-select option { background: #fff; color: var(--text); }
.inline-form { margin: 0; }
.receipt p { margin: 0 0 10px; }
.pagination { justify-content: space-between; align-items: center; margin-top: 16px; }
.page-numbers { display: flex; gap: 6px; flex-wrap: wrap; }
.page-number { min-width: 38px; justify-content: center; }
.page-number.active { background: var(--gray); pointer-events: none; }

.compact-controls {
    padding: 12px;
    margin-bottom: 12px;
    border-top: 0;
}
.compact-toolbar {
    display: flex;
    align-items: center;
    gap: 8px;
    flex-wrap: wrap;
}
.compact-search {
    flex: 1 1 280px;
    min-width: 180px;
}
.compact-search input {
    padding: 8px 10px;
    font-size: 13px;
}
.compact-toolbar .button,
.compact-toolbar button {
    padding: 7px 10px;
    font-size: 13px;
    border-radius: 8px;
}
.compact-filter-label {
    color: var(--muted);
    font-size: 12px;
    font-weight: 700;
    margin-left: 6px;
}
</style>
"""

NAV = """
<div class="navbar">
  <div class="navbar-inner">
    <div class="nav-brand"><div class="georgia-flag"><span class="dot-1"></span><span class="dot-2"></span><span class="dot-3"></span><span class="dot-4"></span></div></div>
    <a class="nav-link" href="/">შეკვეთების მენეჯმენტი</a>
    <a class="nav-link" href="/product-orders">საქონლის შეკვეთები</a>
    <a class="nav-link" href="/sales">გაყიდვები</a>
    <a class="nav-link" href="/excel">Excel ექსპორტი</a>
  </div>
</div>
<script>
(function () {
    let currentVersion = null;
    async function checkForUpdates() {
        try {
            const response = await fetch('/version', { cache: 'no-store' });
            if (!response.ok) return;
            const data = await response.json();
            if (currentVersion === null) {
                currentVersion = data.version;
                return;
            }
            if (data.version !== currentVersion) {
                window.location.reload();
            }
        } catch (error) {
            // Ignore temporary network errors and try again on the next interval.
        }
    }
    checkForUpdates();
    setInterval(checkForUpdates, 3000);
})();
</script>
"""

DEVICE_OPTIONS = """
  <option value="PC">
  <option value="Lenovo">
  <option value="Dell">
  <option value="HP">
  <option value="ASUS">
  <option value="Acer">
  <option value="Apple MacBook">
  <option value="MSI">
  <option value="Samsung">
  <option value="Huawei">
  <option value="Printer HP">
  <option value="Printer Canon">
  <option value="Printer Epson">
"""

PROBLEM_OPTIONS = """
  <option value="Windows">
  <option value="Armushaobs">
  <option value="Chawera">
"""

TEMPLATE_INDEX = CSS + NAV + """
<div class="page">
<h1>შეკვეთების მენეჯმენტი</h1>

<div class="panel compact-controls">
  <form method="get" class="compact-toolbar">
      <div class="compact-search">
        <input name="q" placeholder="ძიება..." value="{{q}}">
      </div>
      <button class="button">ძიება</button>
      <a class="button button-gray" href="/">გასუფთავება</a>
      <span class="compact-filter-label">სტატუსი:</span>
      <a class="button button-status-accepted" href="/?status=მიღებულია">მიღებულია</a>
      <a class="button button-status-work" href="/?status=მიმდინარეობს">მიმდინარეობს</a>
      <a class="button button-status-ready" href="/?status=მზადაა">მზადაა</a>
      <a class="button button-status-given" href="/?status=გაცემული">გაცემული</a>
      <a class="button button-gray" href="/">ყველა</a>
  </form>
</div>

<div class="actions" style="margin-bottom: 16px;">
  <a class="button" href="/new">+ ახალი შეკვეთა</a>
</div>

<div class="table-wrap">
<table class="orders-table">
<tr><th>ID</th><th>კლიენტი</th><th>ტელეფონი</th><th>მოწყობილობა</th><th>პრობლემა</th><th>სტატუსი</th><th>ფასი</th><th>თარიღი და დრო</th><th></th></tr>
{% for r in rows %}
<tr>
<td class="id-cell">{{r[0]}}</td><td>{{r[1]}}</td><td class="phone-cell">{{r[2]}}</td><td class="device-cell">{{r[3]}}</td><td class="problem-cell">{{r[4]}}</td>
<td>
<form class="inline-form" method="post" action="/status/{{r[0]}}">
<input type="hidden" name="next" value="{{current_url}}">
<select class="status-select {% if r[6] == 'მიღებულია' %}status-select-accepted{% elif r[6] == 'მიმდინარეობს' %}status-select-work{% elif r[6] == 'მზადაა' %}status-select-ready{% else %}status-select-given{% endif %}" name="status" onchange="this.form.submit()">
<option value="მიღებულია" {% if r[6] == 'მიღებულია' %}selected{% endif %}>მიღებულია</option>
<option value="მიმდინარეობს" {% if r[6] == 'მიმდინარეობს' %}selected{% endif %}>მიმდინარეობს</option>
<option value="მზადაა" {% if r[6] == 'მზადაა' %}selected{% endif %}>მზადაა</option>
<option value="გაცემული" {% if r[6] == 'გაცემული' %}selected{% endif %}>გაცემული</option>
</select>
</form>
</td>
<td class="price-cell">{{r[7]}}</td>
<td class="date-cell">{{r[8]}}</td>
<td class="open-cell"><a class="button" href="/edit/{{r[0]}}">გახსნა</a></td>
</tr>
{% endfor %}
</table>
</div>

<div class="actions pagination">
  <div class="actions">
    {% if prev_url %}<a class="button button-gray" href="{{prev_url}}">← უკან</a>{% endif %}
    <div class="page-numbers">
      {% for p in page_numbers %}
        <a class="button page-number {% if p == page %}active{% endif %}" href="{{page_urls[p]}}">{{p}}</a>
      {% endfor %}
    </div>
    {% if next_url %}<a class="button" href="{{next_url}}">შემდეგი →</a>{% endif %}
  </div>
</div>
</div>
"""

TEMPLATE_PRODUCT_ORDERS = CSS + NAV + """
<div class="page">
<h1>საქონლის შეკვეთები</h1>
<div class="panel">
  <p class="section-title">შეკვეთის დამატება</p>
  <form method="post" class="grid">
    <div class="col-4"><label>კლიენტი</label><input name="client" required></div>
    <div class="col-4"><label>ტელეფონი</label><input name="phone" type="tel" inputmode="numeric" pattern="[0-9]*" autocomplete="tel" oninput="this.value = this.value.replace(/[^0-9]/g, '')" required></div>
    <div class="col-4"><label>რა შეუკვეთეს</label><input name="item" required></div>
    <div class="col-3"><label>ფასი</label><input name="price" type="tel" inputmode="numeric" pattern="[0-9]*" oninput="this.value = this.value.replace(/[^0-9]/g, '')"></div>
    <div class="col-3"><label>მოსვლის თარიღი</label><input type="date" name="expected_date" required></div>
    <div class="col-6" style="display:flex;align-items:end;"><button type="submit">შეკვეთის დამატება</button></div>
  </form>
</div>
<div class="table-wrap">
<table>
<tr><th>ID</th><th>კლიენტი</th><th>ტელეფონი</th><th>რა შეუკვეთეს</th><th>ფასი</th><th>შეკვეთის თარიღი</th><th>მოსვლის თარიღი</th><th>სტატუსი</th></tr>
{% for o in product_orders %}
<tr>
<td>{{o[0]}}</td>
<td>{{o[1]}}</td>
<td>{{o[2]}}</td>
<td>{{o[3]}}</td>
<td>{{o[4]}} ₾</td>
<td>{{o[5]}}</td>
<td>{{o[6]}}</td>
<td>
<form class="inline-form" method="post" action="/product-order-status/{{o[0]}}">
<select class="status-select {% if o[7] == 'მიღებულია' %}status-select-accepted{% elif o[7] == 'შეკვეთილია' %}status-select-work{% elif o[7] == 'მოსულია' %}status-select-ready{% else %}status-select-given{% endif %}" name="status" onchange="this.form.submit()">
<option value="მიღებულია" {% if o[7] == 'მიღებულია' %}selected{% endif %}>მიღებულია</option>
<option value="შეკვეთილია" {% if o[7] == 'შეკვეთილია' %}selected{% endif %}>შეკვეთილია</option>
<option value="მოსულია" {% if o[7] == 'მოსულია' %}selected{% endif %}>მოსულია</option>
<option value="გაცემული" {% if o[7] == 'გაცემული' %}selected{% endif %}>გაცემული</option>
</select>
</form>
</td>
</tr>
{% endfor %}
</table>
</div>
</div>
"""

TEMPLATE_EXCEL = CSS + NAV + """
<div class="page">
<h1>Excel ექსპორტი</h1>
<div class="panel">
  <p class="section-title">სწრაფი ექსპორტი</p>
  <div class="actions">
      <a class="button" href="/export?period=day">დღე</a>
      <a class="button" href="/export?period=week">კვირა</a>
      <a class="button" href="/export?period=month">თვე</a>
      <a class="button" href="/export?period=all">ყველა</a>
  </div>
</div>
<div class="panel">
  <p class="section-title">დიაპაზონით ექსპორტი</p>
  <form method="get" action="/export" class="grid">
      <input type="hidden" name="period" value="custom">
      <div class="col-3"><label>თარიღიდან</label><input class="date-compact" type="date" name="start_date" required></div>
      <div class="col-3"><label>თარიღამდე</label><input class="date-compact" type="date" name="end_date" required></div>
      <div class="col-6" style="display:flex;align-items:end;"><button class="button" type="submit">ექსპორტი დიაპაზონით</button></div>
  </form>
</div>
</div>
"""

TEMPLATE_NEW = CSS + NAV + """
<div class="page">
<h1>ახალი შეკვეთა</h1>
<div class="panel">
<form method="post" class="grid">
<div class="col-6"><label>კლიენტის სახელი</label><input name="name"></div>
<div class="col-6"><label>ტელეფონი</label><input name="phone" type="tel" inputmode="numeric" pattern="[0-9]*" autocomplete="tel" oninput="this.value = this.value.replace(/[^0-9]/g, '')"></div>
<div class="col-12"><label>მოწყობილობა</label><input name="device" list="deviceTemplatesNew"><datalist id="deviceTemplatesNew">{{device_options|safe}}</datalist></div>
<div class="col-6"><label>პრობლემა</label><input name="problem" list="problemTemplatesNew"><datalist id="problemTemplatesNew">{{problem_options|safe}}</datalist></div>
<div class="col-6"><label>რაც გაკეთდა</label><textarea name="work"></textarea></div>
<div class="col-3"><label>ფასი</label><input name="price" type="tel" inputmode="numeric" pattern="[0-9]*" oninput="this.value = this.value.replace(/[^0-9]/g, '')"></div>
<div class="col-12 actions"><button type="submit">შენახვა</button><a class="button button-gray" href="/">უკან</a></div>
</form>
</div></div>
"""

TEMPLATE_EDIT = CSS + NAV + """
<div class="page">
<h1>შეკვეთის რედაქტირება №{{row[0]}}</h1>
<div class="panel">
<form method="post" class="grid">
<div class="col-6"><label>კლიენტის სახელი</label><input name="name" value="{{row[1]}}"></div>
<div class="col-6"><label>ტელეფონი</label><input name="phone" type="tel" inputmode="numeric" pattern="[0-9]*" autocomplete="tel" oninput="this.value = this.value.replace(/[^0-9]/g, '')" value="{{row[2]}}"></div>
<div class="col-12"><label>მოწყობილობა</label><input name="device" list="deviceTemplatesEdit" value="{{row[3]}}"><datalist id="deviceTemplatesEdit">{{device_options|safe}}</datalist></div>
<div class="col-6"><label>პრობლემა</label><input name="problem" list="problemTemplatesEdit" value="{{row[4]}}"><datalist id="problemTemplatesEdit">{{problem_options|safe}}</datalist></div>
<div class="col-6"><label>რაც გაკეთდა</label><textarea name="work">{{row[5]}}</textarea></div>
<div class="col-3"><label>ფასი</label><input name="price" type="tel" inputmode="numeric" pattern="[0-9]*" oninput="this.value = this.value.replace(/[^0-9]/g, '')" value="{{row[7]}}"></div>
<div class="col-12">
<label>შეცვალე სტატუსი ერთ ღილაკზე დაჭერით</label>
<div class="actions">
<button type="submit" name="status" value="მიღებულია" class="button-status-accepted">მიღებულია</button>
<button type="submit" name="status" value="მიმდინარეობს" class="button-status-work">მიმდინარეობს</button>
<button type="submit" name="status" value="მზადაა" class="button-status-ready">მზადაა</button>
<button type="submit" name="status" value="გაცემული" class="button-status-given">გაცემული</button>
</div>
</div>
<div class="col-12 actions"><a class="button button-red" href="/delete/{{row[0]}}">წაშლა</a><a class="button" href="/print/{{row[0]}}">ქვითარი</a><a class="button button-gray" href="/">უკან</a></div>
</form></div></div>
"""

TEMPLATE_PRINT = CSS + NAV + """
<div class="page">
<h1>ქვითარი №{{row[0]}}</h1>
<div class="panel receipt">
<p><b>თარიღი:</b> {{row[8]}}</p>
<p><b>კლიენტი:</b> {{row[1]}}</p>
<p><b>ტელეფონი:</b> {{row[2]}}</p>
<p><b>მოწყობილობა:</b> {{row[3]}}</p>
<p><b>პრობლემა:</b> {{row[4]}}</p>
<p><b>რაც გაკეთდა:</b> {{row[5]}}</p>
<p><b>სტატუსი:</b> {{row[6]}}</p>
<p><b>ფასი:</b> {{row[7]}} ₾</p>
<div class="actions" style="margin-top: 16px;"><button class="button" onclick="window.print()">ბეჭვა</button><a class="button button-gray" href="/edit/{{row[0]}}">უკან</a></div>
</div>
</div>
"""

TEMPLATE_SALES = CSS + NAV + """
<div class="page">
<h1>გაყიდვები</h1>
<div class="panel">
  <p class="section-title">გაყიდვის დამატება</p>
  <form method="post" class="grid">
    <div class="col-6"><label>საქონლის დასახელება</label><input name="item" required></div>
    <div class="col-3"><label>ფასი</label><input name="price" type="tel" inputmode="numeric" pattern="[0-9]*" oninput="this.value = this.value.replace(/[^0-9]/g, '')" required></div>
    <div class="col-3" style="display:flex;align-items:end;"><button type="submit">დამატება</button></div>
  </form>
</div>
<div class="table-wrap">
<table>
<tr><th>ID</th><th>საქონელი</th><th>ფასი</th><th>თარიღი და დრო</th></tr>
{% for s in sales %}
<tr><td>{{s[0]}}</td><td>{{s[1]}}</td><td>{{s[2]}} ₾</td><td>{{s[3]}}</td></tr>
{% endfor %}
</table>
</div>
<div class="actions pagination">
  <div class="actions">
    {% if prev_url %}<a class="button button-gray" href="{{prev_url}}">← უკან</a>{% endif %}
    <div class="page-numbers">
      {% for p in page_numbers %}
        <a class="button page-number {% if p == page %}active{% endif %}" href="{{page_urls[p]}}">{{p}}</a>
      {% endfor %}
    </div>
    {% if next_url %}<a class="button" href="{{next_url}}">შემდეგი →</a>{% endif %}
  </div>
</div>
</div>
"""


def db():
    return sqlite3.connect('orders.sqlite')


def init_db():
    con = db()
    con.execute('''CREATE TABLE IF NOT EXISTS orders(id INTEGER PRIMARY KEY AUTOINCREMENT,name TEXT,phone TEXT,device TEXT,problem TEXT,work TEXT,status TEXT,price TEXT,date TEXT)''')
    con.execute('''CREATE TABLE IF NOT EXISTS sales(id INTEGER PRIMARY KEY AUTOINCREMENT,item TEXT,price TEXT,date TEXT)''')
    con.execute('''CREATE TABLE IF NOT EXISTS product_orders(id INTEGER PRIMARY KEY AUTOINCREMENT,client TEXT,phone TEXT,item TEXT,price TEXT,order_date TEXT,expected_date TEXT,status TEXT)''')
    con.execute("UPDATE product_orders SET status='მიღებულია' WHERE status='принято'")
    con.execute("UPDATE product_orders SET status='შეკვეთილია' WHERE status='заказано'")
    con.execute("UPDATE product_orders SET status='მოსულია' WHERE status='пришло'")
    con.execute("UPDATE product_orders SET status='გაცემული' WHERE status='выдано'")
    con.execute('''CREATE TABLE IF NOT EXISTS app_state(key TEXT PRIMARY KEY, value INTEGER NOT NULL)''')
    con.execute('INSERT OR IGNORE INTO app_state(key, value) VALUES (?, ?)', ('change_version', 0))
    con.commit()
    con.close()


def get_change_version():
    con = db()
    version = con.execute('SELECT value FROM app_state WHERE key=?', ('change_version',)).fetchone()[0]
    con.close()
    return version


def bump_change_version(con):
    con.execute('UPDATE app_state SET value = value + 1 WHERE key=?', ('change_version',))

def orders_page_url(page, q='', status=''):
    params = {'page': page}
    if q:
        params['q'] = q
    if status:
        params['status'] = status
    return f"/?{urlencode(params)}"


def sales_page_url(page):
    return f"/sales?{urlencode({'page': page})}"

def digits_only(value):
    return ''.join(ch for ch in value if ch.isdigit())


def build_excel(rows, title):
    wb = Workbook(); ws = wb.active; ws.title = 'Orders'
    ws.append(['ID','Client','Phone','Device','Problem','Work','Status','Price','Date'])
    for c in ws[1]: c.font = Font(bold=True)
    for row in rows: ws.append(list(row))
    ws['K1'] = 'Report'; ws['K2'] = title
    stream = BytesIO(); wb.save(stream); stream.seek(0); return stream


@app.route('/version')
def version():
    return jsonify(version=get_change_version())


@app.route('/')
def index():
    q = request.args.get('q', '')
    status = request.args.get('status', '')
    requested_page = max(request.args.get('page', 1, type=int), 1)
    con = db()
    where = []
    params = []
    if q:
        search = f'%{q}%'
        where.append('(name LIKE ? OR phone LIKE ? OR device LIKE ? OR problem LIKE ?)')
        params.extend([search, search, search, search])
    if status:
        where.append('status = ?')
        params.append(status)
    where_sql = f"WHERE {' AND '.join(where)}" if where else ''
    total = con.execute(f'SELECT COUNT(*) FROM orders {where_sql}', params).fetchone()[0]
    last_page = max((total + PER_PAGE - 1) // PER_PAGE, 1)
    page = min(requested_page, last_page)
    offset = (page - 1) * PER_PAGE
    rows = con.execute(
        f'SELECT * FROM orders {where_sql} ORDER BY id DESC LIMIT ? OFFSET ?',
        params + [PER_PAGE, offset],
    ).fetchall()
    con.close()
    first_page = max(page - 2, 1)
    final_page = min(page + 2, last_page)
    page_numbers = list(range(first_page, final_page + 1))
    page_urls = {p: orders_page_url(p, q, status) for p in page_numbers}
    return render_template_string(
        TEMPLATE_INDEX,
        rows=rows,
        q=q,
        status=status,
        page=page,
        total=total,
        last_page=last_page,
        page_numbers=page_numbers,
        page_urls=page_urls,
        prev_url=orders_page_url(page - 1, q, status) if page > 1 else '',
        next_url=orders_page_url(page + 1, q, status) if page < last_page else '',
        current_url=request.full_path if request.query_string else request.path,
    )


@app.route('/product-orders', methods=['GET', 'POST'])
def product_orders():
    con = db()
    if request.method == 'POST':
        con.execute(
            '''INSERT INTO product_orders(client, phone, item, price, order_date, expected_date, status) VALUES (?, ?, ?, ?, ?, ?, ?)''',
            (
                request.form['client'],
                digits_only(request.form['phone']),
                request.form['item'],
                digits_only(request.form['price']),
                datetime.datetime.now().strftime('%Y-%m-%d %H:%M'),
                request.form['expected_date'],
                'მიღებულია',
            ),
        )
        bump_change_version(con)
        con.commit()
        con.close()
        return redirect('/product-orders')
    rows = con.execute('SELECT * FROM product_orders ORDER BY id DESC').fetchall()
    con.close()
    return render_template_string(TEMPLATE_PRODUCT_ORDERS, product_orders=rows)


@app.route('/product-order-status/<id>', methods=['POST'])
def product_order_status(id):
    con = db()
    con.execute('UPDATE product_orders SET status=? WHERE id=?', (request.form['status'], id))
    bump_change_version(con)
    con.commit()
    con.close()
    return redirect('/product-orders')


@app.route('/excel')
def excel_page():
    return render_template_string(TEMPLATE_EXCEL)


@app.route('/sales', methods=['GET', 'POST'])
def sales():
    con = db()
    if request.method == 'POST':
        con.execute(
            'INSERT INTO sales(item, price, date) VALUES (?, ?, ?)',
            (request.form['item'], digits_only(request.form['price']), datetime.datetime.now().strftime('%Y-%m-%d %H:%M')),
        )
        bump_change_version(con)
        con.commit()
        con.close()
        return redirect('/sales')
    requested_page = max(request.args.get('page', 1, type=int), 1)
    total = con.execute('SELECT COUNT(*) FROM sales').fetchone()[0]
    last_page = max((total + PER_PAGE - 1) // PER_PAGE, 1)
    page = min(requested_page, last_page)
    offset = (page - 1) * PER_PAGE
    sales_rows = con.execute('SELECT * FROM sales ORDER BY id DESC LIMIT ? OFFSET ?', (PER_PAGE, offset)).fetchall()
    con.close()
    first_page = max(page - 2, 1)
    final_page = min(page + 2, last_page)
    page_numbers = list(range(first_page, final_page + 1))
    page_urls = {p: sales_page_url(p) for p in page_numbers}
    return render_template_string(
        TEMPLATE_SALES,
        sales=sales_rows,
        page=page,
        page_numbers=page_numbers,
        page_urls=page_urls,
        prev_url=sales_page_url(page - 1) if page > 1 else '',
        next_url=sales_page_url(page + 1) if page < last_page else '',
    )


@app.route('/new', methods=['GET','POST'])
def new():
    if request.method == 'POST':
        con = db()
        con.execute('''INSERT INTO orders(name, phone, device, problem, work, status, price, date) VALUES (?, ?, ?, ?, ?, ?, ?, ?)''', (request.form['name'], digits_only(request.form['phone']), request.form['device'], request.form['problem'], request.form['work'], 'მიღებულია', digits_only(request.form['price']), datetime.datetime.now().strftime('%Y-%m-%d %H:%M')))
        bump_change_version(con)
        con.commit(); con.close(); return redirect('/')
    return render_template_string(TEMPLATE_NEW, device_options=DEVICE_OPTIONS, problem_options=PROBLEM_OPTIONS)


@app.route('/status/<id>', methods=['POST'])
def update_status(id):
    con = db()
    con.execute('UPDATE orders SET status=? WHERE id=?', (request.form['status'], id))
    bump_change_version(con)
    con.commit()
    con.close()
    return redirect(request.form.get('next') or '/')


@app.route('/edit/<id>', methods=['GET','POST'])
def edit(id):
    con = db()
    if request.method == 'POST':
        con.execute('''UPDATE orders SET name=?, phone=?, device=?, problem=?, work=?, status=?, price=? WHERE id=?''', (request.form['name'], digits_only(request.form['phone']), request.form['device'], request.form['problem'], request.form['work'], request.form['status'], digits_only(request.form['price']), id))
        bump_change_version(con)
        con.commit(); con.close(); return redirect('/')
    row = con.execute('SELECT * FROM orders WHERE id=?', (id,)).fetchone()
    con.close()
    return render_template_string(TEMPLATE_EDIT, row=row, device_options=DEVICE_OPTIONS, problem_options=PROBLEM_OPTIONS)


@app.route('/delete/<id>')
def delete(id):
    con = db(); con.execute('DELETE FROM orders WHERE id=?', (id,)); bump_change_version(con); con.commit(); con.close(); return redirect('/')


@app.route('/print/<id>')
def print_order(id):
    con = db(); row = con.execute('SELECT * FROM orders WHERE id=?', (id,)).fetchone(); con.close(); return render_template_string(TEMPLATE_PRINT,row=row)


@app.route('/export')
def export_orders():
    period = request.args.get('period','all'); start_date = request.args.get('start_date',''); end_date = request.args.get('end_date',''); now = datetime.datetime.now(); con = db(); query = 'SELECT * FROM orders'; params = (); report_title = 'All period'
    if period == 'day': d1 = now.strftime('%Y-%m-%d'); d2 = (now + datetime.timedelta(days=1)).strftime('%Y-%m-%d'); query = 'SELECT * FROM orders WHERE date >= ? AND date < ? ORDER BY id DESC'; params = (d1,d2); report_title = f'Day: {d1}'
    elif period == 'week': start = (now - datetime.timedelta(days=7)).strftime('%Y-%m-%d'); end = (now + datetime.timedelta(days=1)).strftime('%Y-%m-%d'); query = 'SELECT * FROM orders WHERE date >= ? AND date < ? ORDER BY id DESC'; params = (start,end); report_title = f'Week: {start} - {end}'
    elif period == 'month': start = (now - datetime.timedelta(days=30)).strftime('%Y-%m-%d'); end = (now + datetime.timedelta(days=1)).strftime('%Y-%m-%d'); query = 'SELECT * FROM orders WHERE date >= ? AND date < ? ORDER BY id DESC'; params = (start,end); report_title = f'Month: {start} - {end}'
    elif period == 'custom' and start_date and end_date: end_plus = (datetime.datetime.strptime(end_date,'%Y-%m-%d') + datetime.timedelta(days=1)).strftime('%Y-%m-%d'); query = 'SELECT * FROM orders WHERE date >= ? AND date < ? ORDER BY id DESC'; params = (start_date,end_plus); report_title = f'Custom: {start_date} - {end_date}'
    else: query = 'SELECT * FROM orders ORDER BY id DESC'
    rows = con.execute(query,params).fetchall(); con.close(); file_obj = build_excel(rows,report_title); safe_title = report_title.replace(' ','_').replace(':','')
    return app.response_class(file_obj.getvalue(), mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet', headers={'Content-Disposition': f'attachment; filename=orders_{safe_title}.xlsx'})


init_db()

if __name__ == '__main__':
    app.run(host='0.0.0.0', port=5000, debug=False)
